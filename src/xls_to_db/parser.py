# -*- coding: utf-8 -*-
from __future__ import absolute_import

import logging
import os
import unicodedata
from StringIO import StringIO
from collections import namedtuple
from datetime import date, datetime, time
from decimal import Decimal
from types import NoneType

import pyexcel as p
from xlrd import XLRDError

from .exceptions import InvalidFieldNameError, InsertError, UnsupportedFileError
from .utils import import_by_name
from .validators import is_integer

logger = logging.getLogger(__name__)


def clean(v, encoding='ascii'):
    return unicodedata.normalize('NFKD', unicode(v)).encode(encoding, 'ignore')


def normalize_table_name(sheet_name):
    ret = clean(sheet_name).lower()
    for char in [' ', '-']:
        ret = ret.replace(char, '_')
    return ret


def normalize_field_name(value):
    ret = clean(value).lower()
    for char in [' ', '-']:
        ret = ret.replace(char, '_')
    return ret


def validate_field_name(value):
    if is_integer(str(value)):
        raise InvalidFieldNameError(value)


class Wrapper(object):
    def __init__(self, filename, file_content=None):
        name, extension = os.path.splitext(filename)
        self.formats = {}
        if extension == '.xls':
            from xlrd import open_workbook
            wb = open_workbook(filename=filename,
                               file_contents=file_content,
                               formatting_info=True)
            for i in range(wb.nsheets):
                sheet = wb.sheet_by_index(i)
                # table_name = normalize_table_name(sheet.name)
                self.formats[i] = {}
                for c in range(sheet.ncols):
                    try:
                        cell = sheet.cell(1, c)
                        xs = wb.xf_list[cell.xf_index]
                        f = wb.format_map[xs.format_key].format_str
                    except IndexError:
                        f = ""

                    self.formats[i][c] = f
        elif extension == '.xlsx':
            from openpyxl import load_workbook
            if file_content:
                wb = load_workbook(StringIO(file_content))
            else:
                wb = load_workbook(filename)
            for i, name in enumerate(wb.sheetnames):
                sheet = wb.get_sheet_by_name(name)
                self.formats[i] = {}
                for col, cells in enumerate(sheet.columns):
                    cell = cells[1]
                    self.formats[i][col] = cell.number_format
        else:
            pass  # pragma: no cover


ORDER = (None, NoneType, int, long, float, Decimal,
         date, time, datetime, str, unicode)


def inspect_column(values, fmt):
    data_type = str
    cast = lambda x: x
    try:
        data_type = type(values[0])
        if data_type in (int,):
            if any([e in fmt for e in ["#,", ".0"]]):
                data_type = Decimal
        elif data_type == date:
            if any([e in fmt for e in ["hh", "mm", "M", "H"]]):
                data_type = datetime
    except IndexError:
        pass
    for value in values:
        current = type(value)
        if ORDER.index(current) > ORDER.index(data_type):
            data_type = current

    return data_type, cast


AAA = namedtuple("AAA", "name,type,cast,format,min_value,max_value,max_length")


def _min(serie):
    if not serie:
        return None
    return min(serie)


def _max(serie):
    if not serie:
        return None
    return max(serie)


class Parser(object):
    def __init__(self, filename, driver="sqlite3", file_content=None,
                 analyze_rows=100, headers=True, prefix=""):
        try:
            if file_content:
                name, extension = os.path.splitext(filename)
                self.book = p.get_book(file_type=extension[1:],
                                       file_content=file_content)
                self.F = Wrapper(filename, file_content=file_content)
            else:
                self.book = p.get_book(file_name=filename)
                self.F = Wrapper(filename, None)
        except XLRDError as e:
            raise UnsupportedFileError("Error reading `{}`: {}".format(filename, e))
        self.analyze_rows = analyze_rows
        self.headers = headers
        self.driver = import_by_name("xls_to_db.drivers.%s.driver" % driver)

        self._names = []
        self._field_defs = None
        self.struct = None
        self.host = None
        self.dbname = ":memory:"
        self.username = ""
        self.password = ""
        self.prefix = prefix
        self.selection_full = self.selection = range(0, self.book.number_of_sheets())

        self.varchar_min = 0
        self.varchar_empty = 10

    @property
    def selection(self):
        return self.__selection

    @selection.setter
    def selection(self, value):
        if not set(value).issubset(self.selection_full):
            raise ValueError("Invalid selection")
        self.__selection = value

    @property
    def table_names(self):
        if not self._names:
            self.analyze()
        return self._names

    def analyze(self):
        self.struct = []
        for i in range(self.book.number_of_sheets()):
            self.struct.append([])
            sheet = self.book.sheet_by_index(i)
            self._names.append(normalize_table_name("{}{}".format(self.prefix, sheet.name)))
            for col_idx in range(sheet.number_of_columns()):
                col = sheet.column_at(col_idx)
                values = col[1:self.analyze_rows + 1] or [0]
                max_length = len(max([clean(v) for v in col], key=len))
                max_value = _min(col[1:])
                min_value = _max(col[1:])
                n = normalize_field_name(col[0] or u"field_%s" % i)
                fmt = self.F.formats[i][col_idx]
                t, c = inspect_column(values, fmt)
                self.struct[-1].append(AAA(n, t, c,
                                           fmt, min_value, max_value,
                                           max_length))
        return self.struct

    @property
    def field_defs(self):
        if not self._field_defs:
            if not self.struct:
                self.analyze()
            self._field_defs = []
            for i, sheet_idx in enumerate(self.selection):
                sheet = self.struct[sheet_idx]
                self._field_defs.append([])
                sheet_field_defs = []
                fields = []
                cfg = {'varchar_min': self.varchar_min,
                       'varchar_empty': self.varchar_empty}
                for col_idx, defs in enumerate(sheet):
                    try:
                        name = defs.name
                        validate_field_name(name)
                        if name in self.driver.RESERVED:
                            name = "%s_%s" % (name, col_idx)
                        fields.append(name)
                        sheet_field_defs.append("{} {}".format(name,
                                                               self.driver.db_type(defs, cfg)))
                    except InvalidFieldNameError as e:
                        raise InvalidFieldNameError(e.field_name, col_idx)
                self._field_defs[i] = sheet_field_defs
        return self._field_defs

    def get_clauses(self):
        clauses = []
        for i, c in enumerate(self.selection):
            defs = self.field_defs[i]
            if defs:
                sql = self.driver.get_create(self.table_names[c], defs)
                clauses.append(sql)

        return clauses

    def get_sql(self):
        c = self.get_clauses()
        return ";".join(c)

    def set_connection(self, host, dbname, username, password):
        self.driver.set_params(host, dbname, username, password)

    def _doit(self, sql, fetch=False):
        return self.driver.execute(sql, fetch=fetch)

    def create_table(self, out=None, skip_if_exists=False):
        stms = self.get_clauses()
        for i, c in enumerate(self.selection):
            if skip_if_exists and self.driver.exists(self.table_names[c]):
                continue
            self.driver.execute(stms[i], fetch=False)
        return stms

    def drop(self):
        for i in self.selection:
            if self.driver.exists(self.table_names[i]):
                sql = 'DROP TABLE %s;' % self.table_names[i]
                self.driver.execute(sql)

    def dump(self):
        for i in self.selection:
            sql = 'SELECT * FROM %s;' % self.table_names[i]
            res = self._doit(sql, fetch=True)
            yield res

    def truncate(self):
        for i in self.selection:
            table = self.table_names[i]
            self.driver.truncate(table)

    def update(self, pkcol):
        # if not self.field_defs:
        #     self.get_fielddefs()
        for i in self.selection:
            try:
                table = self.table_names[i]
                sheet = self.book.sheet_by_index(i)
                pk_name = self.field_defs[i][pkcol].split()[0]
                for row_idx in range(sheet.number_of_rows()):
                    if row_idx == 0:
                        continue
                    row = sheet.row_at(row_idx)
                    values = []
                    for cc, defs in enumerate(self.field_defs[i]):
                        values.append([defs.split()[0], row[cc]])

                    self.driver.update(table, values, [(pk_name, row[pkcol])])
            except Exception:
                raise
            return

    def append(self):
        for i in self.selection:
            try:
                sheet = self.book.sheet_by_index(i)
                table = self.table_names[i]
                params = ['%s'] * sheet.number_of_columns()
                insert = r"INSERT INTO {table} VALUES ({params});".format(table=table,
                                                                          params=", ".join(params))
                for row_idx in range(sheet.number_of_rows()):
                    if row_idx == 0:
                        continue
                    try:
                        row = sheet.row_at(row_idx)
                        self.driver.execute(insert, row, fetch=False)
                    except Exception as e:
                        msg = """Error insert line {} of sheet {} ({})
Error: {}
Stm: {}
Values: {}
""".format(row_idx, i, self.book.sheet_by_index(i).name, e, insert, row)
                        raise InsertError(msg)
                        #     c.execute(insert, row)
                        # conn.commit()
            except Exception as e:
                raise
            return

    load = append
